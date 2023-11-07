# -*- coding: utf-8 -*-
"""
@Time    : 2023/05/14 16:23
@Author  : itlubber
@Site    : itlubber.art
"""

import re
import os
import json
import joblib
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from openpyxl.cell.cell import Cell
from openpyxl.drawing.image import Image
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import NamedStyle, Border, Side, Alignment, PatternFill, Font


def save_pickle(obj, file):
    joblib.dump(obj, file)


class ExcelWriter:

    def __init__(self, style_excel=os.path.join(os.path.dirname(os.path.abspath(__file__)), 'template.xlsx'), style_sheet_name="初始化", mode="replace", fontsize=10, font='楷体', theme_color='2639E9', opacity=0.85):
        """
        excel 文件内容写入公共方法

        :param style_excel: 样式模版文件，默认当前路径下的 template.xlsx ，如果项目路径调整需要进行相应的调整
        :param style_sheet_name: 模版文件内初始样式sheet名称，默认即可
        :param fontsize: 插入excel文件中内容的字体大小，默认 10
        :param font: 插入excel文件中内容的字体，默认 楷体
        :param theme_color: 主题色，默认 2639E9，注意不包含 #
        :param opacity: 写入dataframe时使用颜色填充主题色的透明度设置，默认 0.85
        """
        # english_width，chinese_width
        self.english_width = 0.12
        self.chinese_width = 0.21
        self.mode = mode
        self.font = font
        self.opacity = opacity
        self.fontsize = fontsize
        self.theme_color = theme_color
        self.workbook = load_workbook(style_excel)
        self.style_sheet = self.workbook[style_sheet_name]

        self.name_styles = []
        self.init_style(font, fontsize, theme_color)
        for style in self.name_styles:
            if style.name not in self.workbook.style_names:
                self.workbook.add_named_style(style)

    def add_conditional_formatting(self, worksheet, start_space, end_space):
        """
        设置条件格式

        :param worksheet: 当前选择设置条件格式的sheet
        :param start_space: 开始单元格位置
        :param end_space: 结束单元格位置
        """
        worksheet.conditional_formatting.add(f'{start_space}:{end_space}', DataBarRule(start_type='min', end_type='max', color=self.theme_color))

    @staticmethod
    def set_column_width(worksheet, column, width):
        """
        调整excel列宽

        :param worksheet: 当前选择调整列宽的sheet
        :param column: 列，可以直接输入 index 或者 字母
        :param width: 设置列的宽度
        """
        worksheet.column_dimensions[column if isinstance(column, str) else get_column_letter(column)] = width

    @staticmethod
    def set_number_format(worksheet, space, _format):
        """
        设置数值显示格式

        :param worksheet: 当前选择调整数值显示格式的sheet
        :param space: 单元格范围
        :param _format: 显示格式，参考 openpyxl
        """
        cells = worksheet[space]
        if isinstance(cells, Cell):
            cells = [cells]

        for cell in cells:
            if isinstance(cell, tuple):
                for c in cell:
                    c.number_format = _format
            else:
                cell.number_format = _format

    def get_sheet_by_name(self, name):
        """
        获取sheet名称为name的工作簿，如果不存在，则从初始模版文件中拷贝一个名称为name的sheet

        :param name: 需要获取的工作簿名称
        """
        if name not in self.workbook.sheetnames:
            worksheet = self.workbook.copy_worksheet(self.style_sheet)
            worksheet.title = name
        else:
            worksheet = self.workbook[name]

        return worksheet

    def insert_value2sheet(self, worksheet, insert_space, value="", style="content", auto_width=False):
        """
        向sheet中的某个单元格插入某种样式的内容

        :param worksheet: 需要插入内容的sheet
        :param insert_space: 内容插入的单元格位置，可以是 "B2" 或者 (2, 2) 任意一种形式
        :param value: 需要插入的内容
        :param style: 渲染的样式，参考 init_style 中初始设置的样式
        :param auto_width: 是否开启自动调整列宽
        :return 返回插入元素最后一列之后、最后一行之后的位置
        """
        if isinstance(insert_space, str):
            worksheet[insert_space] = value
            cell = worksheet[insert_space]
            start_col = re.findall('\D+', insert_space)[0]
            start_row = int(re.findall("\d+", insert_space)[0])
        else:
            cell = worksheet.cell(insert_space[0], insert_space[1], value)
            start_col = get_column_letter(insert_space[1])
            start_row = insert_space[0]
        cell.style =  style

        if auto_width:
            curr_width = worksheet.column_dimensions[start_col].width
            auto_width = min(max([(self.check_contain_chinese(value)[1] * self.english_width + self.check_contain_chinese(value)[2] * self.chinese_width) * self.fontsize, 10, curr_width]), 50)
            worksheet.column_dimensions[start_col].width = auto_width

        return start_row + 1, column_index_from_string(start_col) + 1

    def insert_pic2sheet(self, worksheet, fig, insert_space, figsize=(600, 250)):
        """
        向excel中插入图片内容

        :param worksheet: 需要插入内容的sheet
        :param fig: 需要插入的图片路径
        :param insert_space: 插入图片的起始单元格
        :param figsize: 图片大小设置
        :return 返回插入元素最后一列之后、最后一行之后的位置
        """
        if isinstance(insert_space, str):
            start_row = int(re.findall("\d+", insert_space)[0])
            start_col = re.findall('\D+', insert_space)[0]
        else:
            start_row, start_col = insert_space
            start_col = get_column_letter(start_col)

        image = Image(fig)
        image.width, image.height = figsize
        worksheet.add_image(image, f"{start_col}{start_row}")

        return start_row + int(figsize[1] / 17.5), column_index_from_string(start_col) + 8

    def insert_rows(self, worksheet, row, row_index, col_index, merge_rows=None, style="", auto_width=False, style_only=False):
        curr_col = column_index_from_string(col_index)
        for j, v in enumerate(row):
            if merge_rows is not None and row_index + 1 not in merge_rows:
                if j == 0:
                    self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + j)}{row_index}', self.astype_insertvalue(v), style="merge_left", auto_width=auto_width)
                elif j == len(row) - 1:
                    self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + j)}{row_index}', self.astype_insertvalue(v), style="merge_right", auto_width=auto_width)
                else:
                    self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + j)}{row_index}', self.astype_insertvalue(v), style="merge_middle", auto_width=auto_width)
            elif style_only or len(row) <= 1:
                self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + j)}{row_index}', self.astype_insertvalue(v), style=style or "middle", auto_width=auto_width)
            else:
                if j == 0:
                    self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + j)}{row_index}', self.astype_insertvalue(v), style=f"{style}_left" if style else "left", auto_width=auto_width)
                elif j == len(row) - 1:
                    self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + j)}{row_index}', self.astype_insertvalue(v), style=f"{style}_right" if style else "right", auto_width=auto_width)
                else:
                    self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + j)}{row_index}', self.astype_insertvalue(v), style=f"{style}_middle" if style else "middle", auto_width=auto_width)

    def insert_df2sheet(self, worksheet, data, insert_space, merge_column=None, header=True, index=False, auto_width=False, fill=False, merge=False):
        """
        向excel文件中插入指定样式的dataframe数据

        :param worksheet: 需要插入内容的sheet
        :param data: 需要插入的dataframe
        :param insert_space: 插入内容的起始单元格位置
        :param merge_column: 需要分组显示的列，index或者列名
        :param header: 是否存储dataframe的header，暂不支持多级表头
        :param index: 是否存储dataframe的index
        :param auto_width: 是否自动调整列宽
        :param fill: 是否使用颜色填充而非边框
        :param merge: 是否合并单元格，配合 merge_column 一起使用，当前版本仅在 merge_column 只有一列时有效
        
        返回插入元素最后一列之后、最后一行之后的位置
        """
        df = data.copy()

        if isinstance(insert_space, str):
            start_row = int(re.findall("\d+", insert_space)[0])
            start_col = re.findall('\D+', insert_space)[0]
        else:
            start_row, start_col = insert_space
            start_col = get_column_letter(start_col)

        if merge_column:
            if not isinstance(merge_column, (list, np.ndarray)):
                merge_column = [merge_column]

            if isinstance(merge_column[0], (int, float)):
                merge_column = [df.columns.tolist()[col] if col not in df.columns else col for col in merge_column]
                
            df = df.sort_values(merge_column).reset_index(drop=True)
            
            merge_cols = [get_column_letter(df.columns.get_loc(col) + column_index_from_string(start_col)) for col in merge_column]
            merge_rows = list(np.cumsum(df.groupby(merge_column)[merge_column].count().values[:, 0]) + start_row + 1)
        else:
            merge_cols = None
            merge_rows = None

        for i, row in enumerate(dataframe_to_rows(df, header=header, index=index)):
            if fill:
                if i == 0:
                    if header:
                        self.insert_rows(worksheet, row, start_row + i, start_col, style="header", auto_width=auto_width)
                    else:
                        self.insert_rows(worksheet, row, start_row + i, start_col, style="middle_even_first", auto_width=auto_width, style_only=True)
                else:
                    if i % 2 == 1:
                        style = "middle_odd_last" if (header and i == len(df)) or (not header and i + 1 == len(df)) else "middle_odd"
                    else:
                        style = "middle_even_last" if (header and i == len(df)) or (not header and i + 1 == len(df)) else "middle_even"
                    
                    self.insert_rows(worksheet, row, start_row + i, start_col, style=style, auto_width=auto_width, style_only=True)
            else:
                if i == 0:
                    if header:
                        self.insert_rows(worksheet, row, start_row + i, start_col, style="header", auto_width=auto_width)
                    else:
                        self.insert_rows(worksheet, row, start_row + i, start_col, style="first", auto_width=auto_width)
                elif (header and i == len(df)) or (not header and i + 1 == len(df)):
                    self.insert_rows(worksheet, row, start_row + i, start_col, style="last", auto_width=auto_width)
                else:
                    self.insert_rows(worksheet, row, start_row + i, start_col, auto_width=auto_width, merge_rows=merge_rows)

        # 合并单元格, 仅支持单列, 两列及其以上不进行合并
        if merge and merge_column and merge_cols and len(merge_cols) == 1:
            if header:
                merge_rows = [start_row + 2] + merge_rows
            else:
                merge_rows = [r - 1 for r in [start_row + 2] + merge_rows]
            
            for s, e in zip(merge_rows[:-1], merge_rows[1:]):
                if e - s > 1:
                    for merge_col in merge_cols:
                        worksheet.merge_cells(f"{merge_col}{s-1}:{merge_col}{e-1}")

        end_row = start_row + len(data) + 1 if header else start_row + len(data)

        return (end_row, column_index_from_string(start_col) + len(data.columns))

    @staticmethod
    def check_contain_chinese(check_str):
        out = []
        for ch in str(check_str).encode('utf-8').decode('utf-8'):
            if u'\u4e00' <= ch <= u'\u9fff':
                out.append(True)
            else:
                out.append(False)
        return out, len(out) - sum(out), sum(out)

    @staticmethod
    def astype_insertvalue(value, decimal_point=4):
        if re.search('tuple|list|set|numpy.ndarray|Categorical|numpy.dtype|Interval', str(type(value))):
            return str(value)
        elif re.search('float', str(type(value))):
            return round(float(value), decimal_point)
        else:
            return value

    @staticmethod
    def calc_continuous_cnt(list_, index_=0):
        """
        Clac continuous_cnt
        
        Examples:
            list_ = ['A','A','A','A','B','C','C','D','D','D']
            (1) calc_continuous_cnt(list_, 0) ===>('A', 0, 4)
            (2) calc_continuous_cnt(list_, 4) ===>('B', 4, 1)
            (3) calc_continuous_cnt(list_, 6) ===>('C', 6, 1)
        """
        if index_ >= len(list_):
            return None, None, None

        else:
            cnt, str_ = 0, list_[index_]
            for i in range(index_, len(list_), 1):
                if list_[i] == str_:
                    cnt = cnt + 1
                else:
                    break
            return str_, index_, cnt

    @staticmethod
    def itlubber_border(border, color):
        if len(border) == 3:
            return Border(left=Side(border_style=border[0], color=color[0]), right=Side(border_style=border[1], color=color[1]), bottom=Side(border_style=border[2], color=color[2]),)
        else:
            return Border(left=Side(border_style=border[0], color=color[0]), right=Side(border_style=border[1], color=color[1]), bottom=Side(border_style=border[2], color=color[2]), top=Side(border_style=border[3], color=color[3]),)

    @staticmethod
    def get_cell_space(space):
        if isinstance(space, str):
            start_row = int(re.findall("\d+", space)[0])
            start_col = re.findall('\D+', space)[0]
            return start_row, column_index_from_string(start_col)
        else:
            start_row = space[0]
            if isinstance(space[1], int):
                start_col = get_column_letter(space[1])
            else:
                start_col = space[1]
            return f"{start_row}{start_col}"
    
    @staticmethod
    def calculate_rgba_color(hex_color, opacity, prefix="#"):
        rgb_color = tuple(int(hex_color.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
        rgba_color = tuple(int((1 - opacity) * c + opacity * 255) for c in rgb_color)
        
        return prefix + '{:02X}{:02X}{:02X}'.format(*rgba_color)

    def init_style(self, font, fontsize, theme_color):
        header_style, header_left_style, header_middle_style, header_right_style = NamedStyle(name="header"), NamedStyle(name="header_left"), NamedStyle(name="header_middle"), NamedStyle(name="header_right")
        last_style, last_left_style, last_middle_style, last_right_style = NamedStyle(name="last"), NamedStyle(name="last_left"), NamedStyle(name="last_middle"), NamedStyle(name="last_right")
        content_style, left_style, middle_style, right_style = NamedStyle(name="content"), NamedStyle(name="left"), NamedStyle(name="middle"), NamedStyle(name="right")
        merge_style, merge_left_style, merge_middle_style, merge_right_style =  NamedStyle(name="merge"), NamedStyle(name="merge_left"), NamedStyle(name="merge_middle"), NamedStyle(name="merge_right")
        first_style, first_left_style, first_middle_style, first_right_style = NamedStyle(name="first"), NamedStyle(name="first_left"), NamedStyle(name="first_middle"), NamedStyle(name="first_right")
        
        header_font = Font(size=fontsize, name=font, color="FFFFFF", bold=True)
        header_fill = PatternFill(fill_type="solid", start_color=theme_color)
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        content_fill = PatternFill(fill_type="solid", start_color="FFFFFF")
        content_font = Font(size=fontsize, name=font, color="000000")
        even_fill = PatternFill(fill_type="solid", start_color=self.calculate_rgba_color(self.theme_color, self.opacity, prefix=""))

        header_style.font, header_left_style.font, header_middle_style.font, header_right_style.font = header_font, header_font, header_font, header_font
        header_style.fill, header_left_style.fill, header_middle_style.fill, header_right_style.fill = header_fill, header_fill, header_fill, header_fill
        header_style.alignment, header_left_style.alignment, header_middle_style.alignment, header_right_style.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True), alignment, alignment, alignment

        header_style.border = self.itlubber_border(["medium", "medium", "medium", "medium"], [theme_color, theme_color, theme_color, theme_color])
        header_left_style.border = self.itlubber_border(["medium", "thin", "medium", "medium"], [theme_color, "FFFFFF", theme_color, theme_color])
        header_middle_style.border = self.itlubber_border(["thin", "thin", "medium", "medium"], ["FFFFFF", "FFFFFF", theme_color, theme_color])
        header_right_style.border = self.itlubber_border(["thin", "medium", "medium", "medium"], ["FFFFFF", theme_color, theme_color, theme_color])

        last_style.font, last_left_style.font, last_middle_style.font, last_right_style.font = content_font, content_font, content_font, content_font
        last_style.fill, last_left_style.fill, last_middle_style.fill, last_right_style.fill = content_fill, content_fill, content_fill, content_fill
        last_style.alignment, last_left_style.alignment, last_middle_style.alignment, last_right_style.alignment = alignment, alignment, alignment, alignment

        last_style.border = self.itlubber_border(["medium", "medium", "medium"], [theme_color, theme_color, theme_color])
        last_left_style.border = self.itlubber_border(["medium", "thin", "medium"], [theme_color, "FFFFFF", theme_color])
        last_middle_style.border = self.itlubber_border(["thin", "thin", "medium"], ["FFFFFF", "FFFFFF", theme_color])
        last_right_style.border = self.itlubber_border(["thin", "medium", "medium"], ["FFFFFF", theme_color, theme_color])

        content_style.font, left_style.font, middle_style.font, right_style.font = content_font, content_font, content_font, content_font
        content_style.fill, left_style.fill, middle_style.fill, right_style.fill = content_fill, content_fill, content_fill, content_fill
        content_style.alignment, left_style.alignment, middle_style.alignment, right_style.alignment = alignment, alignment, alignment, alignment

        content_style.border = self.itlubber_border(["medium", "medium", "thin"], [theme_color, theme_color, theme_color])
        left_style.border = self.itlubber_border(["medium", "thin", "thin"], [theme_color, "FFFFFF", theme_color])
        middle_style.border = self.itlubber_border(["thin", "medium", "thin"], ["FFFFFF", "FFFFFF", theme_color])
        right_style.border = self.itlubber_border(["thin", "medium", "thin"], ["FFFFFF", theme_color, theme_color])

        merge_style.font, merge_left_style.font, merge_middle_style.font, merge_right_style.font = content_font, content_font, content_font, content_font
        merge_style.fill, merge_left_style.fill, merge_middle_style.fill, merge_right_style.fill = content_fill, content_fill, content_fill, content_fill
        merge_style.alignment, merge_left_style.alignment, merge_middle_style.alignment, merge_right_style.alignment = alignment, alignment, alignment, alignment

        merge_style.border = self.itlubber_border(["medium", "medium", "thin"], ["FFFFFF", "FFFFFF", "FFFFFF"])
        merge_left_style.border = self.itlubber_border(["medium", "thin", "thin"], [theme_color, "FFFFFF", "FFFFFF"])
        merge_middle_style.border = self.itlubber_border(["thin", "medium", "thin"], ["FFFFFF", "FFFFFF", "FFFFFF"])
        merge_right_style.border = self.itlubber_border(["thin", "medium", "thin"], ["FFFFFF", theme_color, "FFFFFF"])

        first_style.font, first_left_style.font, first_middle_style.font, first_right_style.font = content_font, content_font, content_font, content_font
        first_style.fill, first_left_style.fill, first_middle_style.fill, first_right_style.fill = content_fill, content_fill, content_fill, content_fill
        first_style.alignment, first_left_style.alignment, first_middle_style.alignment, first_right_style.alignment = alignment, alignment, alignment, alignment

        first_style.border = self.itlubber_border(["medium", "medium", "thin", "medium"], [theme_color, theme_color, theme_color, theme_color])
        first_left_style.border = self.itlubber_border(["medium", "thin", "thin", "medium"], [theme_color, "FFFFFF", theme_color, theme_color])
        first_middle_style.border = self.itlubber_border(["thin", "thin", "thin", "medium"], ["FFFFFF", "FFFFFF", theme_color, theme_color])
        first_right_style.border = self.itlubber_border(["thin", "medium", "thin", "medium"], ["FFFFFF", theme_color, theme_color, theme_color])
        
        middle_odd_style, middle_odd_first_style, middle_odd_last_style = NamedStyle(name="middle_odd"), NamedStyle(name="middle_odd_first"), NamedStyle(name="middle_odd_last")
        middle_even_style, middle_even_first_style, middle_even_last_style = NamedStyle(name="middle_even"), NamedStyle(name="middle_even_first"), NamedStyle(name="middle_even_last")
        
        middle_odd_style.font, middle_odd_first_style.font, middle_odd_last_style.font, middle_even_style.font, middle_even_first_style.font, middle_even_last_style.font = content_font, content_font, content_font, content_font, content_font, content_font
        middle_odd_style.alignment, middle_odd_first_style.alignment, middle_odd_last_style.alignment, middle_even_style.alignment, middle_even_first_style.alignment, middle_even_last_style.alignment = alignment, alignment, alignment, alignment, alignment, alignment
        middle_odd_style.fill, middle_odd_first_style.fill, middle_odd_last_style.fill, middle_even_style.fill, middle_even_first_style.fill, middle_even_last_style.fill = content_fill, content_fill, content_fill, even_fill, even_fill, even_fill
        
        middle_odd_first_style.border = Border(top=Side(border_style="medium", color=self.theme_color))
        middle_odd_last_style.border = Border(bottom=Side(border_style="medium", color=self.theme_color))
        middle_even_first_style.border = Border(top=Side(border_style="medium", color=self.theme_color))
        middle_even_last_style.border = Border(bottom=Side(border_style="medium", color=self.theme_color))

        self.name_styles.extend([
            header_style, header_left_style, header_middle_style, header_right_style,
            last_style, last_left_style, last_middle_style, last_right_style,
            content_style, left_style, middle_style, right_style,
            merge_style, merge_left_style, merge_middle_style, merge_right_style,
            first_style, first_left_style, first_middle_style, first_right_style,
            middle_odd_style, middle_even_first_style, middle_odd_last_style, middle_even_style, middle_odd_first_style, middle_even_last_style,
        ])

    def save(self, filename, close=True):
        """
        保存excel文件

        :param filename: 需要保存 excel 文件的路径
        :param close: 是否需要释放 writer
        """
        if self.style_sheet.title in self.workbook.sheetnames:
            self.workbook.remove(self.style_sheet)
        
        if os.path.exists(filename) and self.mode == "append":
            _workbook = load_workbook(filename)
            
            for _sheet_name in _workbook.sheetnames:
                if _sheet_name not in self.workbook.sheetnames:
                    _worksheet = self.get_sheet_by_name(_sheet_name)
                    
                    for i, row in enumerate(_workbook[_sheet_name].iter_rows()):
                        for j, cell in enumerate(row):
                            _worksheet.cell(row=i + 1, column=j + 1).value = cell.value
                            _worksheet.cell(row=i + 1, column=j + 1).style = cell.style

                            if i == _workbook[_sheet_name].max_row - 1:
                                _worksheet.column_dimensions[get_column_letter(j + 1)].width = _workbook[_sheet_name].column_dimensions[get_column_letter(j + 1)].width

                self.workbook.move_sheet(_worksheet, offset=-len(self.workbook.sheetnames) + 1)
                
            _workbook.close()
            
        self.workbook.save(filename)
        
        if close:
            self.workbook.close()
            
            
def dataframe2excel(data, excel_writer, sheet_name=None, title=None, header=True, theme_color="2639E9", fill=True, percent_cols=None, condition_cols=None, custom_cols=None, custom_format="#,##0", color_cols=None, start_col=2, start_row=2, mode="replace", writer_params={}, **kwargs):
    """
    向excel文件中插入指定样式的dataframe数据

    :param data: 需要保存的dataframe数据，index默认不保存，如果需要保存先 .reset_index().rename(columns={"index": "索引名称"}) 再保存，有部分索引 reset_index 之后是 0 而非 index，根据实际情况进行修改
    :param excel_writer: 需要保存到的 excel 文件路径或者 ExcelWriter
    :param sheet_name: 需要插入内容的sheet，如果是 Worksheet，则直接向 Worksheet 插入数据
    :param title: 是否在dataframe之前的位置插入一个标题
    :param header: 是否存储dataframe的header，暂不支持多级表头
    :param theme_color: 主题色
    :param fill: 是否使用单元个颜色填充样式还是使用边框样式
    :param percent_cols: 需要显示为百分数的列，仅修改显示格式，不更改数值
    :param condition_cols: 需要显示条件格式的列（无边框渐变数据条）
    :param color_cols: 需要显示为条件格式颜色填充的列（单元格填充渐变色）
    :param custom_cols: 需要显示自定义格式的列，与 custom_format 参数搭配使用
    :param custom_format: 显示的自定义格式，与 custom_cols 参数搭配使用，默认 #,##0 ，即显示为有分隔符的整数
    :param start_col: 在excel中的开始列数，默认 2，即第二列开始
    :param start_row: 在excel中的开始行数，默认 2，即第二行开始，如果 title 有值的话，会从 start_row + 2 行开始插入dataframe数据
    :param mode: excel写入的模式，可选 append 和 replace ，默认 replace ，选择 append 时会在已有的excel文件中增加内容，不覆盖原有内容
    :param writer_params: 透传至 ExcelWriter 内的参数
    :param **kwargs: 其他参数，透传至 insert_df2sheet 方法，例如 传入 auto_width=True 会根据内容自动调整列宽
    :return 返回插入元素最后一列之后、最后一行之后的位置
    """
    if isinstance(excel_writer, ExcelWriter):
        writer = excel_writer
    else:
        writer = ExcelWriter(theme_color=theme_color, mode=mode, **writer_params)
    
        # if os.path.exists(excel_writer) and mode == "append":
        #     workbook = load_workbook(excel_writer)
            
        #     for _sheet_name in workbook.sheetnames:
        #         _worksheet = writer.get_sheet_by_name(_sheet_name or "Sheet1")
                
        #         for i, row in enumerate(workbook[_sheet_name].iter_rows()):
        #             for j, cell in enumerate(row):
        #                 _worksheet.cell(row=i + 1, column=j + 1, value=cell.value)
            
        #     workbook.close()
    
    if isinstance(sheet_name, Worksheet):
        worksheet = sheet_name
    else:
        worksheet = writer.get_sheet_by_name(sheet_name or "Sheet1")

    if title:
        start_row, end_col = writer.insert_value2sheet(worksheet, (start_row, start_col), value=title, style="header")
        start_row += 1

    end_row, end_col = writer.insert_df2sheet(worksheet, data, (start_row, start_col), fill=fill, header=header, **kwargs)
    
    if percent_cols:
        for c in [c for c in percent_cols if c in data.columns]:
            conditional_column = get_column_letter(start_col + data.columns.get_loc(c))
            writer.set_number_format(worksheet, f"{conditional_column}{end_row - len(data)}:{conditional_column}{end_row - 1}", "0.00%")
    
    if custom_cols:
        for c in [c for c in custom_cols if c in data.columns]:
            conditional_column = get_column_letter(start_col + data.columns.get_loc(c))
            writer.set_number_format(worksheet, f"{conditional_column}{end_row - len(data)}:{conditional_column}{end_row - 1}", custom_format)
    
    if condition_cols:
        for c in [c for c in condition_cols if c in data.columns]:
            conditional_column = get_column_letter(start_col + data.columns.get_loc(c))
            writer.add_conditional_formatting(worksheet, f'{conditional_column}{end_row - len(data)}', f'{conditional_column}{end_row - 1}')

    if color_cols:
        for c in [c for c in color_cols if c in data.columns]:
            try:
                rule = ColorScaleRule(start_type='num', start_value=data[c].min(), start_color=theme_color, mid_type='num', mid_value=0., mid_color='FFFFFF', end_type='num', end_value=data[c].max(), end_color=theme_color)
                conditional_column = get_column_letter(start_col + data.columns.get_loc(c))
                worksheet.conditional_formatting.add(f"{conditional_column}{end_row - len(data)}:{conditional_column}{end_row - 1}", rule)
            except:
                import traceback
                traceback.print_exc()
    
    if not isinstance(excel_writer, ExcelWriter) and not isinstance(sheet_name, Worksheet):
        writer.save(excel_writer)
    
    return end_row, end_col


if __name__ == "__main__":
    writer = ExcelWriter()
    worksheet = writer.get_sheet_by_name("模型报告")
    end_row, end_col = writer.insert_value2sheet(worksheet, "B2", value="模型报告", style="header")
    end_row, end_col = writer.insert_value2sheet(worksheet, "B3", value="当前模型主要为评分卡模型", style="content", auto_width=True)
    sample = pd.DataFrame(np.concatenate([np.random.random_sample((10, 10)) * 40, np.random.randint(0, 3, (10, 2))], axis=1), columns=[f"B{i}" for i in range(10)] + ["target", "type"])
    end_row, end_col = writer.insert_df2sheet(worksheet, sample, (end_row + 2, column_index_from_string("B")))
    end_row, end_col = writer.insert_df2sheet(worksheet, sample, (end_row + 2, column_index_from_string("B")), fill=True)
    end_row, end_col = writer.insert_df2sheet(worksheet, sample, (end_row + 2, column_index_from_string("B")), fill=True, header=False)
    end_row, end_col = writer.insert_df2sheet(worksheet, sample, (end_row + 2, column_index_from_string("B")), merge_column="target")
    end_row, end_col = writer.insert_df2sheet(worksheet, sample, (end_row + 2, column_index_from_string("B")), merge_column=["target", "type"])
    end_row, end_col = writer.insert_df2sheet(worksheet, sample, (end_row + 2, column_index_from_string("B")), merge_column=[10, 11])
    end_row, end_col = dataframe2excel(sample, writer, sheet_name="模型报告", start_row=end_row + 2, percent_cols=["B2", "B6"], condition_cols=["B3", "B9"], color_cols=["B4"])
    writer.save("测试样例.xlsx")
